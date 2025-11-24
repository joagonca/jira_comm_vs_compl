"""
Table writing utilities for Excel sheets
"""

from typing import Sequence, Union

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from utils import StatusLabel
from .styles import ExcelStyles


class TableWriter:
    """Handles writing data tables to Excel worksheets"""
    
    def __init__(self, styles: ExcelStyles):
        self.styles = styles
    
    def write_data_table(self, ws: Worksheet, start_row: int, 
                        data: Sequence[Sequence[Union[str, int, float, StatusLabel]]], 
                        apply_status_coloring: bool = False) -> int:
        """Write a data table with headers and return next available row"""
        for row_idx, row_data in enumerate(data):
            current_row = start_row + row_idx
            
            for col_idx, value in enumerate(row_data):
                col_letter = get_column_letter(col_idx + 1)
                cell = ws[f'{col_letter}{current_row}']
                
                if isinstance(value, StatusLabel):
                    cell.value = value.value
                else:
                    cell.value = value
                    
                cell.border = self.styles.border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                if row_idx == 0:
                    cell.font = self.styles.header_font
                    cell.fill = self.styles.header_fill
                
                if apply_status_coloring and row_idx > 0 and col_idx == len(row_data) - 1:
                    if value == StatusLabel.GOOD:
                        cell.fill = self.styles.good_fill
                    elif value == StatusLabel.WARNING:
                        cell.fill = self.styles.warning_fill
                    elif value in (StatusLabel.POOR, StatusLabel.AGED):
                        cell.fill = self.styles.bad_fill
        
        return start_row + len(data)
    
    def write_section_header(self, ws: Worksheet, row: int, title: str) -> int:
        """Write a section header and return next row"""
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = self.styles.subheader_font
        cell.fill = self.styles.subheader_fill
        cell.alignment = Alignment(horizontal='left')
        return row + 1
    
    def write_title(self, ws: Worksheet, title: str, row: int = 1) -> int:
        """Write a sheet title and return next row"""
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = self.styles.title_font
        cell.alignment = self.styles.title_alignment
        return row + 2
    
    def autosize_columns(self, ws: Worksheet) -> None:
        """Auto-size all columns based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column or 1)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
