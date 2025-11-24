"""
Commitment vs Delivery trend chart creator
"""

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import LineChart, Reference

from . import BaseChartCreator


class CommitmentDeliveryTrendChart(BaseChartCreator):
    """Creates line chart for commitment vs delivery trend over time"""
    
    def create(self, ws: Worksheet, sorted_months: list, monthly_metrics: dict, start_row: int) -> None:
        """Create the commitment vs delivery trend chart"""
        title = "Commitment vs Delivery Trend"
        description = "Monthly delivery rates over time for issues and story points. Shows if the team is improving or declining."
        
        data_start = self.write_chart_header(ws, start_row, title, description)
        
        ws[f'A{data_start}'] = "Month"
        ws[f'B{data_start}'] = "Issues Delivery Rate (%)"
        ws[f'C{data_start}'] = "Story Points Delivery Rate (%)"
        
        for i, month_key in enumerate(sorted_months, 1):
            metrics = monthly_metrics[month_key]
            total_issues = metrics['delivered'] + metrics['carryover']
            total_sp = metrics['delivered_sp'] + metrics['carryover_sp']
            
            issue_rate = (metrics['delivered'] / total_issues * 100) if total_issues > 0 else 0
            sp_rate = (metrics['delivered_sp'] / total_sp * 100) if total_sp > 0 else 0
            
            ws[f'A{data_start + i}'] = month_key
            ws[f'B{data_start + i}'] = round(issue_rate, 2)
            ws[f'C{data_start + i}'] = round(sp_rate, 2)
        
        chart = LineChart()
        chart.title = title
        chart.style = 10
        chart.y_axis.title = "Delivery Rate (%)"
        chart.x_axis.title = "Month"
        chart.height = 10
        chart.width = 20
        
        data = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + len(sorted_months), max_col=3)
        cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + len(sorted_months))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f'E{start_row}')
