"""
Cycle time distribution chart creator
"""

import numpy
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import BarChart, Reference

from . import BaseChartCreator


class CycleTimeDistributionChart(BaseChartCreator):
    """Creates bar chart for cycle time distribution by issue type"""
    
    def create(self, ws: Worksheet, cycle_time_per_type: dict, start_row: int) -> None:
        """Create the cycle time distribution chart"""
        title = "Cycle Time Distribution"
        description = "Average cycle times by issue type. Identifies outliers and typical delivery times."
        
        data_start = self.write_chart_header(ws, start_row, title, description)
        
        ws[f'A{data_start}'] = "Issue Type"
        ws[f'B{data_start}'] = "Average Cycle Time (days)"
        
        row_idx = 1
        for issue_type, items in cycle_time_per_type.items():
            values = numpy.array([item[1] for item in items])
            avg_seconds = float(numpy.mean(values))
            avg_days = avg_seconds / (60 * 60 * 24)
            
            ws[f'A{data_start + row_idx}'] = issue_type
            ws[f'B{data_start + row_idx}'] = round(avg_days, 2)
            row_idx += 1
        
        chart = BarChart()
        chart.title = title
        chart.type = "col"
        chart.style = 11
        chart.y_axis.title = "Average Cycle Time (days)"
        chart.x_axis.title = "Issue Type"
        chart.height = 10
        chart.width = 20
        
        data = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + len(cycle_time_per_type))
        cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + len(cycle_time_per_type))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f'E{start_row}')
