"""
Monthly stacked bar chart creator
"""

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import BarChart, Reference

from . import BaseChartCreator


class MonthlyStackedBarChart(BaseChartCreator):
    """Creates stacked bar chart for monthly commitment vs delivery"""
    
    def create(self, ws: Worksheet, sorted_months: list, monthly_metrics: dict, start_row: int) -> None:
        """Create the monthly stacked bar chart"""
        title = "Monthly Commitment vs Delivery"
        description = "Each month shows delivered (green) vs carryover (red) issues side-by-side. Very visual for stakeholders."
        
        data_start = self.write_chart_header(ws, start_row, title, description)
        
        ws[f'A{data_start}'] = "Month"
        ws[f'B{data_start}'] = "Delivered Issues"
        ws[f'C{data_start}'] = "Carryover Issues"
        
        for i, month_key in enumerate(sorted_months, 1):
            metrics = monthly_metrics[month_key]
            ws[f'A{data_start + i}'] = month_key
            ws[f'B{data_start + i}'] = metrics['delivered']
            ws[f'C{data_start + i}'] = metrics['carryover']
        
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = title
        chart.style = 13
        chart.y_axis.title = "Number of Issues"
        chart.x_axis.title = "Month"
        chart.height = 10
        chart.width = 20
        
        data = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + len(sorted_months), max_col=3)
        cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + len(sorted_months))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f'E{start_row}')
