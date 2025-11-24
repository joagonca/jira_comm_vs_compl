"""
Rework ratio trend chart creator
"""

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import LineChart, Reference

from . import BaseChartCreator


class ReworkRatioTrendChart(BaseChartCreator):
    """Creates line chart for rework ratio trend over time"""
    
    def create(self, ws: Worksheet, sorted_months: list, monthly_metrics: dict, start_row: int) -> None:
        """Create the rework ratio trend chart"""
        title = "Rework Ratio Trend"
        description = "Monthly rework percentage over time. Helps identify if technical debt is increasing."
        
        data_start = self.write_chart_header(ws, start_row, title, description)
        
        ws[f'A{data_start}'] = "Month"
        ws[f'B{data_start}'] = "Rework Ratio (%)"
        
        for i, month_key in enumerate(sorted_months, 1):
            metrics = monthly_metrics[month_key]
            defect_effort = metrics['effort_per_type'].get("Defect", 0) + metrics['effort_per_type'].get("Bug", 0)
            story_effort = metrics['effort_per_type'].get("Story", 0)
            total_effort = defect_effort + story_effort
            
            rework_ratio = (defect_effort / total_effort * 100) if total_effort > 0 else 0
            
            ws[f'A{data_start + i}'] = month_key
            ws[f'B{data_start + i}'] = round(rework_ratio, 2)
        
        chart = LineChart()
        chart.title = title
        chart.style = 12
        chart.y_axis.title = "Rework Ratio (%)"
        chart.x_axis.title = "Month"
        chart.height = 10
        chart.width = 20
        
        data = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + len(sorted_months))
        cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + len(sorted_months))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f'E{start_row}')
