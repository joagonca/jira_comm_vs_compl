"""
Base chart creator class
"""

from openpyxl.worksheet.worksheet import Worksheet

from ..styles import ExcelStyles


class BaseChartCreator:    
    def __init__(self, styles: ExcelStyles):
        self.styles = styles
    
    def write_chart_header(self, ws: Worksheet, start_row: int, title: str, description: str) -> int:
        """Write chart title and description"""
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = self.styles.subheader_font
        ws[f'A{start_row+1}'] = description
        ws[f'A{start_row+1}'].font = self.styles.description_font
        return start_row + 3
