"""
Story points pie chart creator
"""

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList

from . import BaseChartCreator


class StoryPointsPieChart(BaseChartCreator):
    """Creates pie chart for story points by outcome"""
    
    def create(self, ws: Worksheet, delivered_sp: float, carryover_sp: float, start_row: int) -> None:
        """Create the story points pie chart"""
        title = "Story Points by Outcome"
        description = "Delivered vs Carryover story points. Simple but effective overview."
        
        data_start = self.write_chart_header(ws, start_row, title, description)
        
        ws[f'A{data_start}'] = "Outcome"
        ws[f'B{data_start}'] = "Story Points"
        
        ws[f'A{data_start + 1}'] = "Delivered"
        ws[f'B{data_start + 1}'] = delivered_sp
        ws[f'A{data_start + 2}'] = "Carryover"
        ws[f'B{data_start + 2}'] = carryover_sp
        
        chart = PieChart()
        chart.title = title
        chart.style = 10
        chart.height = 10
        chart.width = 20
        
        data = Reference(ws, min_col=2, min_row=data_start + 1, max_row=data_start + 2)
        labels = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + 2)
        chart.add_data(data)
        chart.set_categories(labels)
        
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = True
        
        ws.add_chart(chart, f'E{start_row}')
