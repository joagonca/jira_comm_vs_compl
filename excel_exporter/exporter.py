"""
Main Excel exporter class - orchestrates all export operations
"""

import os
from typing import Optional, Tuple, Union

import numpy
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font

from state_manager import State
from utils import (seconds_to_pretty, AGING_THRESHOLDS, 
                   PerformanceThreshold, ReworkThreshold, TrendThreshold, StatusLabel, TrendLabel)

from .styles import ExcelStyles
from .table_writer import TableWriter
from .charts.commitment_delivery import CommitmentDeliveryTrendChart
from .charts.rework import ReworkRatioTrendChart
from .charts.cycle_time import CycleTimeDistributionChart
from .charts.stacked_bar import MonthlyStackedBarChart
from .charts.pie import StoryPointsPieChart


class ExcelExporter:
    """Export JIRA metrics to Excel with multiple sheets"""

    def __init__(self, state: State, output_dir: Optional[str] = None):
        self.state = state
        self.output_dir = output_dir or "."
        self.output_path = self._generate_output_path()
        self.workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        self.styles = ExcelStyles()
        self.table_writer = TableWriter(self.styles)
        
        self.commitment_chart = CommitmentDeliveryTrendChart(self.styles)
        self.rework_chart = ReworkRatioTrendChart(self.styles)
        self.cycle_time_chart = CycleTimeDistributionChart(self.styles)
        self.stacked_bar_chart = MonthlyStackedBarChart(self.styles)
        self.pie_chart = StoryPointsPieChart(self.styles)

    def _generate_filename(self) -> str:
        """Generate filename with project key, team ID, and period"""
        project_key = self.state.get_project_key()
        team_id = self.state.get_team_id()
        
        period = ""
        if self.state.monthly_metrics:
            sorted_months = sorted(self.state.monthly_metrics.keys())
            if len(sorted_months) == 1:
                period = sorted_months[0]
            elif len(sorted_months) > 1:
                period = f"{sorted_months[0]}_to_{sorted_months[-1]}"
        
        parts = []
        if project_key:
            parts.append(project_key)
        if team_id:
            sanitized_team_id = team_id.replace(" ", "").replace(",", "_")
            parts.append(f"Team_{sanitized_team_id}")
        if period:
            parts.append(period)
        
        if not parts:
            parts.append("jira_metrics")
        
        return "_".join(parts) + "_metrics.xlsx"

    def _generate_output_path(self) -> str:
        """Generate full output path combining directory and filename"""
        filename = self._generate_filename()
        return os.path.join(self.output_dir, filename)

    def _generate_sheet_title(self, *parts: str) -> str:
        """Generate sheet title with project key, team ID, and additional parts"""
        project_key = self.state.get_project_key()
        team_id = self.state.get_team_id()
        
        title_parts = []
        if project_key:
            title_parts.append(project_key)
        if team_id:
            title_parts.append(f"Team {team_id}")
        title_parts.extend(parts)
        
        return " - ".join(title_parts) if title_parts else "JIRA Team Performance"

    def export(self) -> str:
        """
        Export all metrics to Excel file
        
        Returns:
            Path to the generated Excel file
        """
        self._create_overall_summary_sheet()
        
        if self.state.monthly_metrics:
            sorted_months = sorted(self.state.monthly_metrics.keys())
            for month_key in sorted_months:
                self._create_monthly_sheet(month_key)
        
        if self.state.monthly_metrics and len(self.state.monthly_metrics) >= 2:
            self._create_visualizations_sheet()
        
        self.workbook.save(self.output_path)
        return self.output_path

    def _create_overall_summary_sheet(self) -> None:
        """Create the overall summary sheet"""
        ws = self.workbook.create_sheet("Overall Summary", 0)
        
        title = self._generate_sheet_title("Overall Summary")
        row = self.table_writer.write_title(ws, title)
        
        row = self._write_commitment_delivery_section(
            ws, row,
            self.state.delivered, self.state.carryover,
            self.state.delivered_sp, self.state.carryover_sp,
            show_total_label=True
        )
        row += 1

        if self.state.monthly_metrics and len(self.state.monthly_metrics) >= 2:
            row = self._write_monthly_trends_section(ws, row)
            row += 1
        
        row = self._write_rework_ratio_section(ws, row, self.state.effort_per_type, "Rework Ratio (Overall)")
        row += 1
        
        row = self._write_cycle_time_by_type_section(ws, row)
        row += 1
        row = self._write_cycle_time_by_sp_section(ws, row)
        row += 1
        
        row = self._write_aging_section(ws, row)
        
        self.table_writer.autosize_columns(ws)

    def _write_monthly_trends_section(self, ws: Worksheet, row: int) -> int:
        """Write monthly trends section"""
        row = self.table_writer.write_section_header(ws, row, "Monthly Trends")
        
        sorted_months = sorted(self.state.monthly_metrics.keys())
        ratios_by_issues = []
        ratios_by_sp = []
        
        for month_key in sorted_months:
            metrics = self.state.monthly_metrics[month_key]
            total = metrics['delivered'] + metrics['carryover']
            total_sp_month = metrics['delivered_sp'] + metrics['carryover_sp']
            
            if total > 0:
                ratios_by_issues.append(metrics['delivered'] / total)
                ratios_by_sp.append((metrics['delivered_sp'] / total_sp_month) if total_sp_month > 0 else 0)
        
        issue_trend = self.state.calculate_linear_trend(ratios_by_issues)
        sp_trend = self.state.calculate_linear_trend(ratios_by_sp)
        
        trend_data = [
            ["Metric", "Trend Direction", "Status"],
            ["Commitment vs Delivery (Issues)", self._trend_to_text(issue_trend), 
             self._trend_to_status(issue_trend)],
            ["Commitment vs Delivery (Story Points)", self._trend_to_text(sp_trend), 
             self._trend_to_status(sp_trend)],
        ]
        
        return self.table_writer.write_data_table(ws, row, trend_data, apply_status_coloring=True)

    def _write_cycle_time_by_type_section(self, ws: Worksheet, row: int) -> int:
        """Write cycle time by issue type section"""
        row = self.table_writer.write_section_header(ws, row, "Average Cycle Time by Issue Type")
        
        cycle_data: list[list[Union[str, int]]] = [["Issue Type", "Count", "Average", "Top 1%", "Bottom 1%", "Std Deviation"]]
        
        for issue_type, items in self.state.cycle_time_per_type.items():
            values = numpy.array([item[1] for item in items])
            cycle_data.append([
                issue_type,
                len(items),
                seconds_to_pretty(float(numpy.mean(values))),
                seconds_to_pretty(float(numpy.percentile(values, 99))),
                seconds_to_pretty(float(numpy.percentile(values, 1))),
                seconds_to_pretty(float(numpy.std(values)))
            ])
        
        return self.table_writer.write_data_table(ws, row, cycle_data)

    def _write_cycle_time_by_sp_section(self, ws: Worksheet, row: int) -> int:
        """Write cycle time by story points section"""
        row = self.table_writer.write_section_header(ws, row, "Average Cycle Time by Story Points")
        
        sp_data: list[list[Union[str, int]]] = [["Story Points", "Count", "Average", "Std Deviation"]]
        sorted_sp_keys = sorted(self.state.cycle_time_per_sp.keys(), 
                               key=lambda x: float('inf') if x == -1 else x)
        
        for sp_key in sorted_sp_keys:
            items = self.state.cycle_time_per_sp[sp_key]
            values = numpy.array([item[1] for item in items])
            sp_display = f"{sp_key} SPs" if sp_key != -1 else "No SPs"
            sp_data.append([
                sp_display,
                len(items),
                seconds_to_pretty(float(numpy.mean(values))),
                seconds_to_pretty(float(numpy.std(values)))
            ])
        
        return self.table_writer.write_data_table(ws, row, sp_data)

    def _write_aging_section(self, ws: Worksheet, row: int) -> int:
        """Write work item aging section"""
        row = self.table_writer.write_section_header(ws, row, "Work Item Aging")
        
        if not self.state.aging_items:
            ws[f'A{row}'] = "No items currently in progress"
            ws[f'A{row}'].font = self.styles.italic_font
            return row + 1
        
        aged_items = [item for item in self.state.aging_items if item['is_aged']]
        
        if not aged_items:
            ws[f'A{row}'] = f"No aged items found ({len(self.state.aging_items)} items in progress)"
            ws[f'A{row}'].font = self.styles.italic_font
            return row + 1
        
        aging_data: list[list[Union[str, int, StatusLabel]]] = [["Issue Key", "Type", "Days In Progress", "Threshold", "Status"]]
        
        for item in sorted(aged_items, key=lambda x: x['days'], reverse=True):
            threshold = AGING_THRESHOLDS.get(item['type'], 14)
            aging_data.append([
                item['key'],
                item['type'],
                f"{item['days']:.1f}",
                threshold,
                StatusLabel.AGED if item['is_aged'] else StatusLabel.OK
            ])
        
        return self.table_writer.write_data_table(ws, row, aging_data, apply_status_coloring=True)

    def _create_monthly_sheet(self, month_key: str) -> None:
        """Create a sheet for a specific month"""
        ws = self.workbook.create_sheet(month_key)
        metrics = self.state.monthly_metrics[month_key]
        
        title = self._generate_sheet_title(month_key)
        row = self.table_writer.write_title(ws, title)
        row = self._write_commitment_delivery_section(
            ws, row,
            metrics['delivered'], metrics['carryover'],
            metrics['delivered_sp'], metrics['carryover_sp'],
            show_total_label=False
        )
        row += 1
        row = self._write_rework_ratio_section(ws, row, metrics['effort_per_type'])
        
        self.table_writer.autosize_columns(ws)

    def _calculate_ratios(self, delivered: int, carryover: int, 
                         delivered_sp: float, carryover_sp: float) -> Tuple[float, float]:
        """Calculate commitment vs delivery ratios"""
        total_issues = delivered + carryover
        total_sp = delivered_sp + carryover_sp
        ratio_issue = (delivered / total_issues * 100) if total_issues > 0 else 0
        ratio_sp = (delivered_sp / total_sp * 100) if total_sp > 0 else 0
        return ratio_issue, ratio_sp

    def _write_commitment_delivery_section(self, ws: Worksheet, row: int, delivered: int, carryover: int,
                                          delivered_sp: float, carryover_sp: float,
                                          show_total_label: bool = False) -> int:
        """Write commitment vs delivery section (reusable for overall and monthly)"""
        row = self.table_writer.write_section_header(ws, row, "Commitment vs Delivery")
        
        total_issues = delivered + carryover
        total_sp = delivered_sp + carryover_sp
        ratio_issue, ratio_sp = self._calculate_ratios(delivered, carryover, delivered_sp, carryover_sp)
        
        total_label = "Total Valid Issues" if show_total_label else "Total Issues"
        
        data = [
            ["Metric", "Value", "Percentage", "Status"],
            [total_label, total_issues, "", ""],
            ["Delivered Issues", delivered, f"{ratio_issue:.2f}%", self._get_status(ratio_issue)],
            ["Carryover Issues", carryover, f"{100-ratio_issue:.2f}%", ""],
            ["Total Story Points", total_sp, "", ""],
            ["Delivered Story Points", delivered_sp, f"{ratio_sp:.2f}%", self._get_status(ratio_sp)],
            ["Carryover Story Points", carryover_sp, f"{100-ratio_sp:.2f}%", ""],
        ]
        
        return self.table_writer.write_data_table(ws, row, data, apply_status_coloring=True)

    def _calculate_rework_metrics(self, effort_per_type: dict) -> Tuple[float, float, float]:
        """Calculate rework ratio and related metrics"""
        defect_effort = effort_per_type.get("Defect", 0) + effort_per_type.get("Bug", 0)
        story_effort = effort_per_type.get("Story", 0)
        total_effort = defect_effort + story_effort
        return defect_effort, story_effort, total_effort

    def _write_rework_ratio_section(self, ws: Worksheet, row: int, effort_per_type: dict, 
                                    title: str = "Rework Ratio") -> int:
        """Write rework ratio section (reusable for overall and monthly)"""
        row = self.table_writer.write_section_header(ws, row, title)
        
        defect_effort, story_effort, total_effort = self._calculate_rework_metrics(effort_per_type)
        
        if total_effort > 0:
            rework_ratio = (defect_effort / total_effort) * 100
            rework_data = [
                ["Metric", "Time (seconds)", "Percentage", "Status"],
                ["Fixing Time (Defects + Bugs)", defect_effort, f"{rework_ratio:.2f}%", 
                 self._get_rework_status(rework_ratio)],
                ["Building Time (Stories)", story_effort, f"{100-rework_ratio:.2f}%", ""],
                ["Total Time", total_effort, "100.00%", ""],
            ]
        else:
            rework_data = [
                ["Metric", "Value"],
                ["Status", "No data available"],
            ]
        
        return self.table_writer.write_data_table(ws, row, rework_data, apply_status_coloring=True)

    def _get_status(self, percentage: float) -> StatusLabel:
        """Get status label based on percentage (for commitment/delivery)"""
        if percentage >= PerformanceThreshold.GOOD:
            return StatusLabel.GOOD
        elif percentage >= PerformanceThreshold.WARNING:
            return StatusLabel.WARNING
        else:
            return StatusLabel.POOR

    def _get_rework_status(self, percentage: float) -> StatusLabel:
        """Get status label for rework percentage (inverted - lower is better)"""
        if percentage >= ReworkThreshold.POOR:
            return StatusLabel.POOR
        elif percentage >= ReworkThreshold.WARNING:
            return StatusLabel.WARNING
        else:
            return StatusLabel.GOOD

    def _trend_to_text(self, slope: float, threshold: float = TrendThreshold.COMMITMENT_DELIVERY) -> str:
        """Convert trend slope to text description"""
        if slope > threshold:
            return TrendLabel.IMPROVING.value
        elif slope < -threshold:
            return TrendLabel.DECLINING.value
        else:
            return TrendLabel.STABLE.value

    def _trend_to_status(self, slope: float, threshold: float = TrendThreshold.COMMITMENT_DELIVERY) -> StatusLabel:
        """Convert trend slope to status for coloring"""
        if slope > threshold:
            return StatusLabel.GOOD
        elif slope < -threshold:
            return StatusLabel.POOR
        else:
            return StatusLabel.WARNING

    def _create_visualizations_sheet(self) -> None:
        """Create visualizations sheet with charts for overall metrics"""
        ws = self.workbook.create_sheet("Visualizations")
        
        title = self._generate_sheet_title("Visualizations")
        ws.merge_cells('A1:P1')
        cell = ws['A1']
        cell.value = title
        cell.font = self.styles.title_font
        cell.alignment = self.styles.title_alignment
        
        sorted_months = sorted(self.state.monthly_metrics.keys())
        
        self.commitment_chart.create(ws, sorted_months, self.state.monthly_metrics, start_row=3)
        self.rework_chart.create(ws, sorted_months, self.state.monthly_metrics, start_row=20)
        self.cycle_time_chart.create(ws, self.state.cycle_time_per_type, start_row=37)
        self.stacked_bar_chart.create(ws, sorted_months, self.state.monthly_metrics, start_row=54)
        self.pie_chart.create(ws, self.state.delivered_sp, self.state.carryover_sp, start_row=71)
