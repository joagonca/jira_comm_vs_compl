"""
Excel styling and formatting utilities
"""

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ExcelStyles:
    """Centralized styles for Excel workbook"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        
        self.subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        self.subheader_font = Font(bold=True, size=11)
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        self.title_font = Font(bold=True, size=14)
        self.title_alignment = Alignment(horizontal='center')
        
        self.description_font = Font(italic=True, size=9)
        
        self.italic_font = Font(italic=True)
