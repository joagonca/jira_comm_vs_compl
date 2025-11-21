"""
Excel exporter for JIRA metrics
"""

import os
from typing import Optional, Tuple, Union, Sequence

import numpy
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

from state_manager import State
from utils import (seconds_to_pretty, AGING_THRESHOLDS, 
                   PerformanceThreshold, ReworkThreshold, TrendThreshold, StatusLabel, TrendLabel)


class ExcelExporter:
    """Export JIRA metrics to Excel with multiple sheets"""

    def __init__(self, state: State, output_dir: Optional[str] = None):
        self.state = state
        self.output_dir = output_dir or "."
        self.output_path = self._generate_output_path()
        self.workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
        
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        self.subheader_font = Font(bold=True, size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        self.bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def _generate_filename(self) -> str:
        """Generate filename with project key and period"""
        project_key = self.state.get_project_key()
        
        # Get period from monthly metrics
        period = ""
        if self.state.monthly_metrics:
            sorted_months = sorted(self.state.monthly_metrics.keys())
            if len(sorted_months) == 1:
                period = sorted_months[0]
            elif len(sorted_months) > 1:
                period = f"{sorted_months[0]}_to_{sorted_months[-1]}"
        
        parts = []
        if project_key:
            parts.append(project_key)
        if period:
            parts.append(period)
        
        if not parts:
            parts.append("jira_metrics")
        
        return "_".join(parts) + "_metrics.xlsx"

    def _generate_output_path(self) -> str:
        """Generate full output path combining directory and filename"""
        filename = self._generate_filename()
        return os.path.join(self.output_dir, filename)

    def _generate_sheet_title(self, *parts: str) -> str:
        """Generate sheet title with project key and additional parts"""
        project_key = self.state.get_project_key()
        
        title_parts = []
        if project_key:
            title_parts.append(project_key)
        title_parts.extend(parts)
        
        return " - ".join(title_parts) if title_parts else "JIRA Team Performance"

    def export(self) -> str:
        """
        Export all metrics to Excel file
        
        Returns:
            Path to the generated Excel file
        """
        self._create_overall_summary_sheet()
        
        if self.state.monthly_metrics:
            sorted_months = sorted(self.state.monthly_metrics.keys())
            for month_key in sorted_months:
                self._create_monthly_sheet(month_key)
        
        if self.state.monthly_metrics and len(self.state.monthly_metrics) >= 2:
            self._create_visualizations_sheet()
        
        self.workbook.save(self.output_path)
        return self.output_path

    def _create_overall_summary_sheet(self) -> None:
        """Create the overall summary sheet"""
        ws = self.workbook.create_sheet("Overall Summary", 0)
        
        title = self._generate_sheet_title("Overall Summary")
        row = self._write_title(ws, title)
        
        row = self._write_commitment_delivery_section(
            ws, row,
            self.state.delivered, self.state.carryover,
            self.state.delivered_sp, self.state.carryover_sp,
            show_total_label=True
        )
        row += 1

        if self.state.monthly_metrics and len(self.state.monthly_metrics) >= 2:
            row = self._write_monthly_trends_section(ws, row)
            row += 1
        
        row = self._write_rework_ratio_section(ws, row, self.state.effort_per_type, "Rework Ratio (Overall)")
        row += 1
        
        row = self._write_cycle_time_by_type_section(ws, row)
        row += 1
        row = self._write_cycle_time_by_sp_section(ws, row)
        row += 1
        
        row = self._write_aging_section(ws, row)
        
        self._autosize_columns(ws)

    def _write_monthly_trends_section(self, ws: Worksheet, row: int) -> int:
        """Write monthly trends section"""
        row = self._write_section_header(ws, row, "Monthly Trends")
        
        sorted_months = sorted(self.state.monthly_metrics.keys())
        ratios_by_issues = []
        ratios_by_sp = []
        
        for month_key in sorted_months:
            metrics = self.state.monthly_metrics[month_key]
            total = metrics['delivered'] + metrics['carryover']
            total_sp_month = metrics['delivered_sp'] + metrics['carryover_sp']
            
            if total > 0:
                ratios_by_issues.append(metrics['delivered'] / total)
                ratios_by_sp.append((metrics['delivered_sp'] / total_sp_month) if total_sp_month > 0 else 0)
        
        issue_trend = self.state.calculate_linear_trend(ratios_by_issues)
        sp_trend = self.state.calculate_linear_trend(ratios_by_sp)
        
        trend_data = [
            ["Metric", "Trend Direction", "Status"],
            ["Commitment vs Delivery (Issues)", self._trend_to_text(issue_trend), 
             self._trend_to_status(issue_trend)],
            ["Commitment vs Delivery (Story Points)", self._trend_to_text(sp_trend), 
             self._trend_to_status(sp_trend)],
        ]
        
        return self._write_data_table(ws, row, trend_data, apply_status_coloring=True)

    def _write_cycle_time_by_type_section(self, ws: Worksheet, row: int) -> int:
        """Write cycle time by issue type section"""
        row = self._write_section_header(ws, row, "Average Cycle Time by Issue Type")
        
        cycle_data: list[list[Union[str, int]]] = [["Issue Type", "Count", "Average", "Top 1%", "Bottom 1%", "Std Deviation"]]
        
        for issue_type, items in self.state.cycle_time_per_type.items():
            values = numpy.array([item[1] for item in items])
            cycle_data.append([
                issue_type,
                len(items),
                seconds_to_pretty(float(numpy.mean(values))),
                seconds_to_pretty(float(numpy.percentile(values, 99))),
                seconds_to_pretty(float(numpy.percentile(values, 1))),
                seconds_to_pretty(float(numpy.std(values)))
            ])
        
        return self._write_data_table(ws, row, cycle_data)

    def _write_cycle_time_by_sp_section(self, ws: Worksheet, row: int) -> int:
        """Write cycle time by story points section"""
        row = self._write_section_header(ws, row, "Average Cycle Time by Story Points")
        
        sp_data: list[list[Union[str, int]]] = [["Story Points", "Count", "Average", "Std Deviation"]]
        sorted_sp_keys = sorted(self.state.cycle_time_per_sp.keys(), 
                               key=lambda x: float('inf') if x == -1 else x)
        
        for sp_key in sorted_sp_keys:
            items = self.state.cycle_time_per_sp[sp_key]
            values = numpy.array([item[1] for item in items])
            sp_display = f"{sp_key} SPs" if sp_key != -1 else "No SPs"
            sp_data.append([
                sp_display,
                len(items),
                seconds_to_pretty(float(numpy.mean(values))),
                seconds_to_pretty(float(numpy.std(values)))
            ])
        
        return self._write_data_table(ws, row, sp_data)

    def _write_aging_section(self, ws: Worksheet, row: int) -> int:
        """Write work item aging section"""
        row = self._write_section_header(ws, row, "Work Item Aging")
        
        if not self.state.aging_items:
            ws[f'A{row}'] = "No items currently in progress"
            ws[f'A{row}'].font = Font(italic=True)
            return row + 1
        
        aged_items = [item for item in self.state.aging_items if item['is_aged']]
        
        if not aged_items:
            ws[f'A{row}'] = f"No aged items found ({len(self.state.aging_items)} items in progress)"
            ws[f'A{row}'].font = Font(italic=True)
            return row + 1
        
        aging_data: list[list[Union[str, int, StatusLabel]]] = [["Issue Key", "Type", "Days In Progress", "Threshold", "Status"]]
        
        for item in sorted(aged_items, key=lambda x: x['days'], reverse=True):
            threshold = AGING_THRESHOLDS.get(item['type'], 14)
            aging_data.append([
                item['key'],
                item['type'],
                f"{item['days']:.1f}",
                threshold,
                StatusLabel.AGED if item['is_aged'] else StatusLabel.OK
            ])
        
        return self._write_data_table(ws, row, aging_data, apply_status_coloring=True)

    def _create_monthly_sheet(self, month_key: str) -> None:
        """Create a sheet for a specific month"""
        ws = self.workbook.create_sheet(month_key)
        metrics = self.state.monthly_metrics[month_key]
        
        title = self._generate_sheet_title(month_key)
        row = self._write_title(ws, title)
        row = self._write_commitment_delivery_section(
            ws, row,
            metrics['delivered'], metrics['carryover'],
            metrics['delivered_sp'], metrics['carryover_sp'],
            show_total_label=False
        )
        row += 1
        row = self._write_rework_ratio_section(ws, row, metrics['effort_per_type'])
        
        self._autosize_columns(ws)

    def _write_title(self, ws: Worksheet, title: str) -> int:
        """Write a sheet title and return next row"""
        row = 1
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')
        return row + 2

    def _calculate_ratios(self, delivered: int, carryover: int, 
                         delivered_sp: float, carryover_sp: float) -> Tuple[float, float]:
        """Calculate commitment vs delivery ratios"""
        total_issues = delivered + carryover
        total_sp = delivered_sp + carryover_sp
        ratio_issue = (delivered / total_issues * 100) if total_issues > 0 else 0
        ratio_sp = (delivered_sp / total_sp * 100) if total_sp > 0 else 0
        return ratio_issue, ratio_sp

    def _write_commitment_delivery_section(self, ws: Worksheet, row: int, delivered: int, carryover: int,
                                          delivered_sp: float, carryover_sp: float,
                                          show_total_label: bool = False) -> int:
        """Write commitment vs delivery section (reusable for overall and monthly)"""
        row = self._write_section_header(ws, row, "Commitment vs Delivery")
        
        total_issues = delivered + carryover
        total_sp = delivered_sp + carryover_sp
        ratio_issue, ratio_sp = self._calculate_ratios(delivered, carryover, delivered_sp, carryover_sp)
        
        total_label = "Total Valid Issues" if show_total_label else "Total Issues"
        
        data = [
            ["Metric", "Value", "Percentage", "Status"],
            [total_label, total_issues, "", ""],
            ["Delivered Issues", delivered, f"{ratio_issue:.2f}%", self._get_status(ratio_issue)],
            ["Carryover Issues", carryover, f"{100-ratio_issue:.2f}%", ""],
            ["Total Story Points", total_sp, "", ""],
            ["Delivered Story Points", delivered_sp, f"{ratio_sp:.2f}%", self._get_status(ratio_sp)],
            ["Carryover Story Points", carryover_sp, f"{100-ratio_sp:.2f}%", ""],
        ]
        
        return self._write_data_table(ws, row, data, apply_status_coloring=True)

    def _calculate_rework_metrics(self, effort_per_type: dict) -> Tuple[float, float, float]:
        """Calculate rework ratio and related metrics"""
        defect_effort = effort_per_type.get("Defect", 0) + effort_per_type.get("Bug", 0)
        story_effort = effort_per_type.get("Story", 0)
        total_effort = defect_effort + story_effort
        return defect_effort, story_effort, total_effort

    def _write_rework_ratio_section(self, ws: Worksheet, row: int, effort_per_type: dict, 
                                    title: str = "Rework Ratio") -> int:
        """Write rework ratio section (reusable for overall and monthly)"""
        row = self._write_section_header(ws, row, title)
        
        defect_effort, story_effort, total_effort = self._calculate_rework_metrics(effort_per_type)
        
        if total_effort > 0:
            rework_ratio = (defect_effort / total_effort) * 100
            rework_data = [
                ["Metric", "Time (seconds)", "Percentage", "Status"],
                ["Fixing Time (Defects + Bugs)", defect_effort, f"{rework_ratio:.2f}%", 
                 self._get_rework_status(rework_ratio)],
                ["Building Time (Stories)", story_effort, f"{100-rework_ratio:.2f}%", ""],
                ["Total Time", total_effort, "100.00%", ""],
            ]
        else:
            rework_data = [
                ["Metric", "Value"],
                ["Status", "No data available"],
            ]
        
        return self._write_data_table(ws, row, rework_data, apply_status_coloring=True)

    def _write_section_header(self, ws: Worksheet, row: int, title: str) -> int:
        """Write a section header and return next row"""
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = title
        cell.font = self.subheader_font
        cell.fill = self.subheader_fill
        cell.alignment = Alignment(horizontal='left')
        return row + 1

    def _write_data_table(self, ws: Worksheet, start_row: int, 
                         data: Sequence[Sequence[Union[str, int, float, StatusLabel]]], 
                         apply_status_coloring: bool = False) -> int:
        """Write a data table with headers and return next available row"""
        for row_idx, row_data in enumerate(data):
            current_row = start_row + row_idx
            
            for col_idx, value in enumerate(row_data):
                col_letter = get_column_letter(col_idx + 1)
                cell = ws[f'{col_letter}{current_row}']
                
                if isinstance(value, StatusLabel):
                    cell.value = value.value
                else:
                    cell.value = value
                    
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                if row_idx == 0:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                
                if apply_status_coloring and row_idx > 0 and col_idx == len(row_data) - 1:
                    if value == StatusLabel.GOOD:
                        cell.fill = self.good_fill
                    elif value == StatusLabel.WARNING:
                        cell.fill = self.warning_fill
                    elif value in (StatusLabel.POOR, StatusLabel.AGED):
                        cell.fill = self.bad_fill
        
        return start_row + len(data)

    def _get_status(self, percentage: float) -> StatusLabel:
        """Get status label based on percentage (for commitment/delivery)"""
        if percentage >= PerformanceThreshold.GOOD:
            return StatusLabel.GOOD
        elif percentage >= PerformanceThreshold.WARNING:
            return StatusLabel.WARNING
        else:
            return StatusLabel.POOR

    def _get_rework_status(self, percentage: float) -> StatusLabel:
        """Get status label for rework percentage (inverted - lower is better)"""
        if percentage >= ReworkThreshold.POOR:
            return StatusLabel.POOR
        elif percentage >= ReworkThreshold.WARNING:
            return StatusLabel.WARNING
        else:
            return StatusLabel.GOOD

    def _trend_to_text(self, slope: float, threshold: float = TrendThreshold.COMMITMENT_DELIVERY) -> str:
        """Convert trend slope to text description"""
        if slope > threshold:
            return TrendLabel.IMPROVING.value
        elif slope < -threshold:
            return TrendLabel.DECLINING.value
        else:
            return TrendLabel.STABLE.value

    def _trend_to_status(self, slope: float, threshold: float = TrendThreshold.COMMITMENT_DELIVERY) -> StatusLabel:
        """Convert trend slope to status for coloring"""
        if slope > threshold:
            return StatusLabel.GOOD
        elif slope < -threshold:
            return StatusLabel.POOR
        else:
            return StatusLabel.WARNING

    def _autosize_columns(self, ws: Worksheet) -> None:
        """Auto-size all columns based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column or 1)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_visualizations_sheet(self) -> None:
        """Create visualizations sheet with charts for overall metrics"""
        ws = self.workbook.create_sheet("Visualizations")
        
        title = self._generate_sheet_title("Visualizations")
        ws.merge_cells('A1:P1')
        cell = ws['A1']
        cell.value = title
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal='center')
        
        sorted_months = sorted(self.state.monthly_metrics.keys())
        self._create_commitment_delivery_trend_chart(ws, sorted_months, start_row=3)
        self._create_rework_ratio_trend_chart(ws, sorted_months, start_row=20)
        self._create_cycle_time_distribution_chart(ws, start_row=37)
        self._create_monthly_stacked_bar_chart(ws, sorted_months, start_row=54)
        self._create_story_points_pie_chart(ws, start_row=71)

    def _create_commitment_delivery_trend_chart(self, ws: Worksheet, sorted_months: list, start_row: int) -> None:
        """Create line chart for commitment vs delivery trend over time"""
        ws[f'A{start_row}'] = "Commitment vs Delivery Trend"
        ws[f'A{start_row}'].font = self.subheader_font
        ws[f'A{start_row+1}'] = "Monthly delivery rates over time for issues and story points. Shows if the team is improving or declining."
        ws[f'A{start_row+1}'].font = Font(italic=True, size=9)
        
        data_start = start_row + 3
        ws[f'A{data_start}'] = "Month"
        ws[f'B{data_start}'] = "Issues Delivery Rate (%)"
        ws[f'C{data_start}'] = "Story Points Delivery Rate (%)"
        
        for i, month_key in enumerate(sorted_months, 1):
            metrics = self.state.monthly_metrics[month_key]
            total_issues = metrics['delivered'] + metrics['carryover']
            total_sp = metrics['delivered_sp'] + metrics['carryover_sp']
            
            issue_rate = (metrics['delivered'] / total_issues * 100) if total_issues > 0 else 0
            sp_rate = (metrics['delivered_sp'] / total_sp * 100) if total_sp > 0 else 0
            
            ws[f'A{data_start + i}'] = month_key
            ws[f'B{data_start + i}'] = round(issue_rate, 2)
            ws[f'C{data_start + i}'] = round(sp_rate, 2)
        
        chart = LineChart()
        chart.title = "Commitment vs Delivery Trend"
        chart.style = 10
        chart.y_axis.title = "Delivery Rate (%)"
        chart.x_axis.title = "Month"
        chart.height = 10
        chart.width = 20
        
        data = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + len(sorted_months), max_col=3)
        cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + len(sorted_months))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f'E{start_row}')

    def _create_rework_ratio_trend_chart(self, ws: Worksheet, sorted_months: list, start_row: int) -> None:
        """Create line chart for rework ratio trend over time"""
        ws[f'A{start_row}'] = "Rework Ratio Trend"
        ws[f'A{start_row}'].font = self.subheader_font
        ws[f'A{start_row+1}'] = "Monthly rework percentage over time. Helps identify if technical debt is increasing."
        ws[f'A{start_row+1}'].font = Font(italic=True, size=9)
        
        data_start = start_row + 3
        ws[f'A{data_start}'] = "Month"
        ws[f'B{data_start}'] = "Rework Ratio (%)"
        
        for i, month_key in enumerate(sorted_months, 1):
            metrics = self.state.monthly_metrics[month_key]
            defect_effort = metrics['effort_per_type'].get("Defect", 0) + metrics['effort_per_type'].get("Bug", 0)
            story_effort = metrics['effort_per_type'].get("Story", 0)
            total_effort = defect_effort + story_effort
            
            rework_ratio = (defect_effort / total_effort * 100) if total_effort > 0 else 0
            
            ws[f'A{data_start + i}'] = month_key
            ws[f'B{data_start + i}'] = round(rework_ratio, 2)
        
        chart = LineChart()
        chart.title = "Rework Ratio Trend"
        chart.style = 12
        chart.y_axis.title = "Rework Ratio (%)"
        chart.x_axis.title = "Month"
        chart.height = 10
        chart.width = 20
        
        data = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + len(sorted_months))
        cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + len(sorted_months))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f'E{start_row}')

    def _create_cycle_time_distribution_chart(self, ws: Worksheet, start_row: int) -> None:
        """Create bar chart for cycle time distribution by issue type"""
        ws[f'A{start_row}'] = "Cycle Time Distribution"
        ws[f'A{start_row}'].font = self.subheader_font
        ws[f'A{start_row+1}'] = "Average cycle times by issue type. Identifies outliers and typical delivery times."
        ws[f'A{start_row+1}'].font = Font(italic=True, size=9)
        
        data_start = start_row + 3
        ws[f'A{data_start}'] = "Issue Type"
        ws[f'B{data_start}'] = "Average Cycle Time"
        
        row_idx = 1
        for issue_type, items in self.state.cycle_time_per_type.items():
            values = numpy.array([item[1] for item in items])
            avg_seconds = float(numpy.mean(values))
            
            ws[f'A{data_start + row_idx}'] = issue_type
            ws[f'B{data_start + row_idx}'] = seconds_to_pretty(avg_seconds)
            row_idx += 1
        
        chart = BarChart()
        chart.title = "Cycle Time Distribution"
        chart.type = "col"
        chart.style = 11
        chart.y_axis.title = "Average Cycle Time"
        chart.x_axis.title = "Issue Type"
        chart.height = 10
        chart.width = 20
        
        data = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + len(self.state.cycle_time_per_type))
        cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + len(self.state.cycle_time_per_type))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f'E{start_row}')

    def _create_monthly_stacked_bar_chart(self, ws: Worksheet, sorted_months: list, start_row: int) -> None:
        """Create stacked bar chart for monthly commitment vs delivery"""
        ws[f'A{start_row}'] = "Monthly Commitment vs Delivery"
        ws[f'A{start_row}'].font = self.subheader_font
        ws[f'A{start_row+1}'] = "Each month shows delivered (green) vs carryover (red) issues side-by-side. Very visual for stakeholders."
        ws[f'A{start_row+1}'].font = Font(italic=True, size=9)
        
        data_start = start_row + 3
        ws[f'A{data_start}'] = "Month"
        ws[f'B{data_start}'] = "Delivered Issues"
        ws[f'C{data_start}'] = "Carryover Issues"
        
        for i, month_key in enumerate(sorted_months, 1):
            metrics = self.state.monthly_metrics[month_key]
            ws[f'A{data_start + i}'] = month_key
            ws[f'B{data_start + i}'] = metrics['delivered']
            ws[f'C{data_start + i}'] = metrics['carryover']
        
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = "Monthly Commitment vs Delivery"
        chart.style = 13
        chart.y_axis.title = "Number of Issues"
        chart.x_axis.title = "Month"
        chart.height = 10
        chart.width = 20
        
        data = Reference(ws, min_col=2, min_row=data_start, max_row=data_start + len(sorted_months), max_col=3)
        cats = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + len(sorted_months))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        chart.series[0].graphicalProperties.solidFill = self.good_fill.start_color.rgb
        chart.series[1].graphicalProperties.solidFill = self.bad_fill.start_color.rgb
        
        ws.add_chart(chart, f'E{start_row}')

    def _create_story_points_pie_chart(self, ws: Worksheet, start_row: int) -> None:
        """Create pie chart for story points by outcome"""
        ws[f'A{start_row}'] = "Story Points by Outcome"
        ws[f'A{start_row}'].font = self.subheader_font
        ws[f'A{start_row+1}'] = "Delivered vs Carryover story points. Simple but effective overview."
        ws[f'A{start_row+1}'].font = Font(italic=True, size=9)
        
        data_start = start_row + 3
        ws[f'A{data_start}'] = "Outcome"
        ws[f'B{data_start}'] = "Story Points"
        
        ws[f'A{data_start + 1}'] = "Delivered"
        ws[f'B{data_start + 1}'] = self.state.delivered_sp
        ws[f'A{data_start + 2}'] = "Carryover"
        ws[f'B{data_start + 2}'] = self.state.carryover_sp
        
        chart = PieChart()
        chart.title = "Story Points by Outcome"
        chart.style = 10
        chart.height = 10
        chart.width = 20
        
        data = Reference(ws, min_col=2, min_row=data_start + 1, max_row=data_start + 2)
        labels = Reference(ws, min_col=1, min_row=data_start + 1, max_row=data_start + 2)
        chart.add_data(data)
        chart.set_categories(labels)
        
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = True
        
        ws.add_chart(chart, f'E{start_row}')
